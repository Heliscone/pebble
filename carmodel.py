import numpy as np
import matplotlib.pyplot as plt
from dataclasses import dataclass
from typing import Optional
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import time
np.set_printoptions(legacy='1.25')

# bump is applied vertically 
# sus points class has: geometry information for a given wheel
# master sus points class has: four sus points
# contact patch class has: contact patch forces/moments/location.
# linkload class has: link loads for a wheel
# corner class has: sus points, contact patch, linkload objects. calculates linkloads from sus points and contactpatch info
# carprops class has: car properties
# master contact patch class has: four contact patches, calculated from carprops and acceleration
# car class has: four corners, carprops, master contact patch. constructs corners from sus, constructs master contact patch from carprops and acceleration
# tire ellipse: generates ellipse points from parameters in gs

class SusPoints:
    '''SusPoints class contains all the geometric information for the suspension points relevant to a single wheel, sorted by inboard and outboard as relevant. It also contains the vectors for suspension members (directed away from the upright).'''
    def __init__(self,lowFore,lowAft,lbj,upFore,upAft,ubj,tieIn,tieOut,pushIn,pushOut,pivotFore,pivotAft,heave,roll):
        self.lowFore = np.array(lowFore)
        self.lowAft = np.array(lowAft)
        self.lbj = np.array(lbj)
        self.upFore = np.array(upFore)
        self.upAft = np.array(upAft)
        self.ubj = np.array(ubj)
        self.tieIn = np.array(tieIn)
        self.tieOut = np.array(tieOut)
        self.pushIn = np.array(pushIn)
        self.pushOut = np.array(pushOut)
        self.pivotFore = np.array(pivotFore)
        self.pivotAft = np.array(pivotAft)
        self.heave = np.array(heave)
        self.roll = np.array(roll)
        self.lowForeVec = self.lowFore - self.lbj
        self.lowAftVec = self.lowAft - self.lbj
        self.upForeVec = self.upFore - self.ubj
        self.upAftVec = self.upAft - self.ubj
        self.tieVec = self.tieIn - self.tieOut
        self.pushVec = self.pushIn - self.pushOut
        self.linkVecList = [self.lowForeVec, self.lowAftVec, self.upForeVec, self.upAftVec, self.tieVec, self.pushVec]
@dataclass
class MasterSusPoints:
    '''MasterSusPoints is a dataclass of four SusPoint objects, one for each wheel.'''
    FL: 'SusPoints'
    FR: 'SusPoints'
    RL: 'SusPoints'
    RR: 'SusPoints'
    name: Optional[str] = None
class ContactPatch:
    '''Contactpatch is a class that contains the coordinates of a given contact patch as well as the forces ON THE WHEEL and the moments about the contact patch ON THE WHEEL'''
    def __init__ (self,forces:np.ndarray,moments:np.ndarray,contact_point_coords:np.ndarray):
        self.contact_patch_forces = forces
        self.contact_patch_moments = moments
        self.contact_point_coords = contact_point_coords
        self.combined_load = np.concatenate((forces,moments))
class LinkLoad:
    '''LinkLoad stores the loads (tension positive) in each of the sus links for a wheel.'''
    def __init__ (self,lowFore:float,lowAft:float,upFore:float,upAft:float,tie:float,push:float):
        self.lowFore = lowFore
        self.lowAft = lowAft
        self.upFore = upFore
        self.upAft = upAft
        self.tie = tie
        self.push = push
        self.loadlist = [lowFore,lowAft,upFore,upAft,tie,push]
class Corner:
    '''Corner objects contain all the information about one corner of the car: sus points, contact patch, and loads. When we create the car, it will consist of four corner objects. We give the corner object suspension points and contact patch information, and it calculates link loads.'''
    def __init__ (self,sus:SusPoints,contact_patch:ContactPatch,link_load:LinkLoad=None):
        self.contact_patch = contact_patch
        self.sus = sus
        # Given the information about contact patch forces, construct link loads
        cp_to_link = self._contact_patch_loads_to_link_loads()
        self.link_load = LinkLoad(* (cp_to_link @ contact_patch.combined_load).flatten())
    def _contact_patch_loads_to_link_loads(self):
        '''Construct a matrix that transforms a 6-vector of contact patch forces and moments (Fx,Fy,Fz,Mx,My,Mz) into a 6-vector of link loads (tensions in lowFore, lowAft, upFore, upAft, tie, push). Keep in mind that the contact_patch.combined_load vector are the forces and moments ON THE WHEEL.'''
        link_vectors = self.sus.linkVecList
        cp_location = self.contact_patch.contact_point_coords
        r_from_contact_patch = [ point - cp_location for point in [self.sus.lbj,self.sus.lbj,self.sus.ubj,self.sus.ubj,self.sus.tieOut,self.sus.pushOut]] # positions of link attachment points relative to contact patch (for moment calculation)
        f_hats = np.array([vec/np.linalg.norm(vec) for vec in link_vectors])
        moment_multipliers = np.array([np.cross(r, f_hat) for r, f_hat in zip(r_from_contact_patch, f_hats)]) # each row is the moment generated by a unit force in the link
        link_loads_to_cp_loads = np.vstack((f_hats.T, moment_multipliers.T)) # 6x6 matrix transforming link loads to contact patch loads
        cp_loads_to_link_loads = -1*np.linalg.inv(link_loads_to_cp_loads)
        # We multiply by -1 because cp.combined_load gives the loads on the wheel, so link_to_cp @ link_loads + cp.combined_load = 0, whence link_loads = -cp_to_link @ cp.combined_load
        # print(cp_loads_to_link_loads)
        return cp_loads_to_link_loads
@dataclass
class CarProps:
    '''CarProps is a dataclass with lots of important car properties.'''
    wheelbase: float
    trackwidth: float
    cg_height: float
    mass: float
    cp_height: float
    cl_a: float
    aero_load: float
    cd_a: float
    drag_force: float
    v: float
    fwb: float
    aero_fwb: float
    gear_ratio: float
    max_amk_torque: float
    tire_radius: float
    LLTD: float
    bump_amount: float
    bump_moment_arm_outward: float
class MasterContactPatch:
    '''MasterContactPatch objects are made with car acceleration and properties, and generates four ContactPatch objects, one for each wheel, by calculating contact patch loads for each wheel using a very simplified load model that says a given LLTD of the lateral load transfer is borne by the front axle. With a conservative sweep of LLTD values, we can get enough of an idea of link load ranges for sizing. finally!'''
    def __init__(self,a_x:float,a_y:float,props:CarProps):
        self.a_x = a_x
        self.a_y = a_y
        self.props = props
        m = props.mass
        g = 9.81 #MY26 is (in most situations) on earth. MYMoon and MYHell haven't yet begun development.
        wheel_center_coord = np.array([props.wheelbase/2, 0, 0])
        hwb,htw = props.wheelbase/2,props.trackwidth/2
        
        LLTD = self.props.LLTD #fraction of lateral load transfer borne by front axle.

        N_fl,N_fr,N_rl,N_rr = self._normal_solver(a_x,a_y,LLTD)
        N_total = m*g + props.aero_load
        max_Fx = props.max_amk_torque*props.gear_ratio/(props.tire_radius)
        forces = [np.array([min((a_x*m+props.drag_force)*N/N_total,max_Fx),a_y*m*N/N_total,-N+props.bump_amount]) for N in [N_fl,N_fr,N_rl,N_rr]]
        bump_moment_arm_RH = np.array([0,props.bump_moment_arm_outward,0])
        bump_moment_arm_LH = np.array([0,-props.bump_moment_arm_outward,0])
        bump_force = np.array([0,0,props.bump_amount])
        RH_bump = np.cross(bump_moment_arm_RH,bump_force)
        LH_bump = np.cross(bump_moment_arm_LH,bump_force)

        cp_offset_list = [[hwb,-htw,0],[hwb,htw,0],[-hwb,-htw,0],[-hwb,htw,0]] 
        cp_list = [np.array(cp) + wheel_center_coord for cp in cp_offset_list] # list of places where we find C.P. (fl,fr,rl,rr as always)

        [fl_cp_coord,fr_cp_coord,rl_cp_coord,rr_cp_coord] = cp_list
        self.flCP = ContactPatch(forces[0],LH_bump,fl_cp_coord)
        self.frCP = ContactPatch(forces[1],RH_bump,fr_cp_coord)
        self.rlCP = ContactPatch(forces[2],LH_bump,rl_cp_coord)
        self.rrCP = ContactPatch(forces[3],RH_bump,rr_cp_coord)
    def _normal_solver(self,a_x,a_y,LLTD):
        '''Given long,lat acceleration and lateral load transfer front bias, return normal loads on each tire.'''
        g = 9.81
        m = self.props.mass
        aero = self.props.aero_load
        fwb = self.props.fwb
        aero_fwb = self.props.aero_fwb
        drag = self.props.drag_force
        h_cp = self.props.cp_height
        h_cg = self.props.cg_height
        wheelbase = self.props.wheelbase
        trackwidth = self.props.trackwidth
        # in these calcs, normals point up
        # F/B balance via bicycle model:
        #   Z axis force balance:
        #       N_f+N_r = m*g + aero
        #   Rear contact patch moment balance:
        #       N_f*wheelbase = m*g*wheelbase*fwb + aero*wheelbase*aero_fwb - drag*h_cp - m*a_x*h_cg
        N_front = (m * g * wheelbase * fwb + aero * wheelbase * aero_fwb - drag * h_cp - m * a_x * h_cg) / wheelbase
        N_rear = m*g + aero - N_front
        # L/R balance:
        #   Total lateral load transfer:
        #       N_l + N_r = m*g+aero
        #   Moment about left contact patch:
        #       N_r*trackwidth = -m*a_y*h_cg + aero*trackwidth/2 + m*g*trackwidth/2
        N_r = (-m*a_y*h_cg + (m*g + aero)*trackwidth/2)/trackwidth
        N_l = m * g + aero - N_r
        LLT = (N_l - N_r)
        # in the positive case, left tire has more load than right
        LLT_f = LLT * LLTD
        LLT_r = LLT * (1 - LLTD)
        N_fl = max((N_front + LLT_f) / 2,0)
        N_fr = max((N_front - LLT_f) / 2,0)
        N_rl = max((N_rear + LLT_r) / 2,0)
        N_rr = max((N_rear - LLT_r) / 2,0)
        return N_fl,N_fr,N_rl,N_rr
class Car:
    '''Quite a big class. Contains all four sus points, properties, and a given acceleratrion. From the acceleration and properties, it constructs a MasterContactPatch, and for each of these contact patches and corresponding sus points, it constructs a Corner object which calculates link loads. Succintish way to go from car + sus + accel to link loads. Hoozah.'''
    def __init__(self,sus:MasterSusPoints,props:CarProps,a_x:float,a_y:float):
        self.wheelbase = props.wheelbase
        self.trackwidth = props.trackwidth
        self.flSus = sus.FL
        self.frSus = sus.FR
        self.rlSus = sus.RL
        self.rrSus = sus.RR
        self.props = props
        self.masterCP = MasterContactPatch(a_x,a_y,props)
        self.flCP = self.masterCP.flCP
        self.frCP = self.masterCP.frCP
        self.rlCP = self.masterCP.rlCP
        self.rrCP = self.masterCP.rrCP
        self.frCorner = Corner(self.frSus,self.frCP)
        # wait for 30 seconds
        self.flCorner = Corner(self.flSus,self.flCP)
        self.rlCorner = Corner(self.rlSus,self.rlCP)
        self.rrCorner = Corner(self.rrSus,self.rrCP)
def tire_ellipse(max_accel_g,max_braking_g,max_lateral_g,num_points):
    '''Generate points a_x,a_y for a tire ellipse in list form.'''
    g = 9.81
    max_accel = max_accel_g * g
    max_braking = max_braking_g * g
    max_lateral = max_lateral_g * g
    angles = np.linspace(0,2*np.pi,num_points,endpoint=False)
    a_x_points = abs(max_braking) * np.cos(angles)
    a_y_points = max_lateral * np.sin(angles)
    ellipse_points = [(a_x,a_y) for a_x,a_y in zip(a_x_points,a_y_points)]
    return ellipse_points
def analyze_link_loads(fl_link_loads,fr_link_loads,rl_link_loads,rr_link_loads):
    '''Return max and min link loads for each corner given time-indexed link loads lists.'''
    def get_max_min(link_loads):
        load_array = np.array([[load.lowFore,load.lowAft,load.upFore,load.upAft,load.tie,load.push] for load in link_loads])
        max_loads = np.max(load_array,axis=0)
        min_loads = np.min(load_array,axis=0)
        return max_loads,min_loads
    fl_max,fl_min = get_max_min(fl_link_loads)
    fr_max,fr_min = get_max_min(fr_link_loads)
    rl_max,rl_min = get_max_min(rl_link_loads)
    rr_max,rr_min = get_max_min(rr_link_loads)
    return (fl_max,fl_min),(fr_max,fr_min),(rl_max,rl_min),(rr_max,rr_min)
def real_sweep_tire_ellipse(tire_ellipse,masterSus,props):
    list_of_cars = []
    for a_x,a_y in tire_ellipse:
        car = Car(sus=masterSus,props=props,a_x=a_x,a_y=a_y)
        list_of_cars.append(car)
    return list_of_cars
def get_link_loads_from_carlist(list_of_cars):
    fl_link_loads = [] #list of LinkLoad objects
    fr_link_loads = []
    rl_link_loads = []
    rr_link_loads = []
    for car in list_of_cars:
        fl_link_loads.append(car.flCorner.link_load.loadlist)
        fr_link_loads.append(car.frCorner.link_load.loadlist)
        rl_link_loads.append(car.rlCorner.link_load.loadlist)
        rr_link_loads.append(car.rrCorner.link_load.loadlist)
    return fl_link_loads,fr_link_loads,rl_link_loads,rr_link_loads
def get_link_loads_from_car(car):
    fl_link_load = car.flCorner.link_load.loadlist
    fr_link_load = car.frCorner.link_load.loadlist
    rl_link_load = car.rlCorner.link_load.loadlist
    rr_link_load = car.rrCorner.link_load.loadlist
    return fl_link_load,fr_link_load,rl_link_load,rr_link_load
def get_extreme_linkload(everything_bagel):
    '''Given dictionary of (LLTD,bump):list_of_cars, return dictionary of (LLTD,bump,ax,ay):(max link loads for each corner) for all points'''
    maxlinkload_dict = {}
    minlinkload_dict = {}
    for ellipse_point_key in everything_bagel:
        ellipse_point_dict= everything_bagel[ellipse_point_key]
        for prop_key in ellipse_point_dict:
            my_car_list = ellipse_point_dict[prop_key]
            fl_link_loads,fr_link_loads,rl_link_loads,rr_link_loads = get_link_loads_from_carlist(my_car_list)
            fr_max_list,fr_min_list = np.max(np.array(fr_link_loads),axis=0).flatten(),np.min(np.array(fr_link_loads),axis=0).flatten()
            fl_max_list,fl_min_list = np.max(np.array(fl_link_loads),axis=0).flatten(),np.min(np.array(fl_link_loads),axis=0).flatten()
            rl_max_list,rl_min_list = np.max(np.array(rl_link_loads),axis=0).flatten(),np.min(np.array(rl_link_loads),axis=0).flatten()
            rr_max_list,rr_min_list = np.max(np.array(rr_link_loads),axis=0).flatten(),np.min(np.array(rr_link_loads),axis=0).flatten()
            maxlinkload_dict[ellipse_point_key] = (fl_max_list,fr_max_list,rl_max_list,rr_max_list,prop_key)
            minlinkload_dict[ellipse_point_key] = (fl_min_list,fr_min_list,rl_min_list,rr_min_list,prop_key)
    return maxlinkload_dict,minlinkload_dict
def get_maxmin_linkload(corner_linkload_dict,selector):
    all_link_forces = np.vstack(list(corner_linkload_dict.values()))
    if selector == "tension":
        return all_link_forces.max(axis=0)
    elif selector == "compression":
        return all_link_forces.min(axis=0)
def get_corner_tensile_extremes(maxlinkload_dict):
    fl_corner_dict = {}
    fr_corner_dict = {}
    rl_corner_dict = {}
    rr_corner_dict = {}
    for key in maxlinkload_dict:
        fl_list,fr_list,rl_list,rr_list,_ = maxlinkload_dict[key]
        fl_corner_dict[key] = fl_list
        fr_corner_dict[key] = fr_list
        rl_corner_dict[key] = rl_list
        rr_corner_dict[key] = rr_list
    fl_max = get_maxmin_linkload(fl_corner_dict,"tension")
    fr_max = get_maxmin_linkload(fr_corner_dict,"tension")
    rl_max = get_maxmin_linkload(rl_corner_dict,"tension")
    rr_max = get_maxmin_linkload(rr_corner_dict,"tension")
    return [fl_max,fr_max,rl_max,rr_max]
def get_corner_compressive_extremes(minlinkload_dict):
    fl_corner_dict = {}
    fr_corner_dict = {}
    rl_corner_dict = {}
    rr_corner_dict = {}
    for key in minlinkload_dict:
        fl_list,fr_list,rl_list,rr_list,_ = minlinkload_dict[key]
        fl_corner_dict[key] = fl_list
        fr_corner_dict[key] = fr_list
        rl_corner_dict[key] = rl_list
        rr_corner_dict[key] = rr_list
    fl_min = get_maxmin_linkload(fl_corner_dict,"compression")
    fr_min = get_maxmin_linkload(fr_corner_dict,"compression")
    rl_min = get_maxmin_linkload(rl_corner_dict,"compression")
    rr_min = get_maxmin_linkload(rr_corner_dict,"compression")
    return [fl_min,fr_min,rl_min,rr_min]
def get_max_thing(name,dataframe):
    return dataframe.loc[dataframe[name].abs().idxmax()][name]

FLsusPointsFromNico = """1658.1	-199.6	-102.0
1362.8	-199.6	-111.1
1536.2	-596.3	-101.0
1648.1	-248.9	-217.5
1372.1	-248.9	-195.2
1517.5	-558.3	-279.7
1600.2	-213.7	-128.3
1628.4	-560.3	-144.2
1517.5	-144.8	-622.7
1517.5	-534.6	-288.9
1517.5	-116.5	-563.6
1459.0	-116.5	-563.6
1517.5	-101.5	-613.5
1459.0	-95.5	-518.8"""
FRsusPointsFromNico = """1658.1	199.6	-102.0
1362.8	199.6	-111.1
1536.2	596.3	-101.0
1648.1	248.9	-217.5
1372.1	248.9	-195.2
1517.5	558.3	-279.7
1600.2	213.7	-128.3
1628.4	560.3	-144.2
1517.5	144.8	-622.7
1517.5	534.6	-288.9
1517.5	116.5	-563.6
1459.0	116.5	-563.6
1517.5	101.5	-613.5
1459.0	95.5	-608.5"""
RLsusPointsFromNico = """26.2	-193.0	-122.4
-164.3	-193.0	-117.1
-17.8	-594.1	-107.3
39.5	-274.3	-223.7
-147.2	-274.3	-239.8
-17.8	-541.8	-289.7
49.5	-294.2	-177.3
76.2	-568.1	-197.8
-17.8	-206.4	-531.1
-17.8	-520.6	-305.0
-74.9	-171.5	-482.6
-17.8	-171.5	-482.6
-17.8	-165.5	-532.1
-74.9	-157.2	-435.2"""
RRsusPointsFromNico = """26.2	193.0	-122.4
-164.3	193.0	-117.1
-17.8	594.1	-107.3
39.5	274.3	-223.7
-147.2	274.3	-239.8
-17.8	541.8	-289.7
49.5	294.2	-177.3
76.2	568.1	-197.8
-17.8	206.4	-531.1
-17.8	520.6	-305.0
-74.9	171.5	-482.6
-17.8	171.5	-482.6
-17.8	165.5	-532.1
-74.9	157.2	-530.0"""
def parse_sus_points(data:str)->SusPoints:
    lines = data.strip().splitlines()
    points = [tuple(float(coord.replace(',',''))/1000 for coord in line.split()) for line in lines]
    return SusPoints(*points)   
flSus = parse_sus_points(FLsusPointsFromNico)
frSus = parse_sus_points(FRsusPointsFromNico)
rlSus = parse_sus_points(RLsusPointsFromNico)
rrSus = parse_sus_points(RRsusPointsFromNico)
masterSus = MasterSusPoints(FL=flSus, FR=frSus, RL=rlSus, RR=rrSus, name="Nico Sus Points")

props = CarProps(
    wheelbase=1.5367, #60.5"
    trackwidth=1.27, #50"
    cg_height=0.0254*11.5, # 11.5" 
    cp_height = 0.36,
    mass=295,
    cl_a = 2.8,
    cd_a=1.2,
    v=60.5,
    fwb=0.48,
    aero_fwb=0.52,
    aero_load=6060,
    drag_force=2589,
    gear_ratio=10.5,
    max_amk_torque=21,
    tire_radius=0.2032,
    LLTD=0.55,
    bump_amount=2000,
    bump_moment_arm_outward=0)
### PROPERTIES. HELLO. CHANGE THESE. ###
max_accel_g = 1.8
max_braking_g = -1.8
max_lateral_g = 1.97
num_ellipse_points=48
pushrod_points = 25
min_LLTD = 0.4
max_LLTD = 0.55
LLTD_points = 5
max_bump = -2000 #N, must be negative
bump_amt_points= 3
bump_location_inch = 3
bump_location_points = 3
min_v=12.5
max_v=30.5
v_points = 5

test_car = Car(sus=masterSus,props=props,a_x=0,a_y=0)
# print("hallo")
print("normal forces: ",test_car.masterCP._normal_solver(a_x=-12.48,a_y=-13.66,LLTD=props.LLTD))

### END PROPERTIES. BYE. ###
sim_property_dp = pd.DataFrame([
    {
        "accel (g)":max_accel_g,
        "braking (g)":max_braking_g,
        "lateral (g)":max_lateral_g,
        "v min (m/s)":min_v,
        "v max (m/s)":max_v,
        "LLTD min":min_LLTD,
        "LLTD max":max_LLTD,
        "bump (N)":max_bump,
        "bump arm (inches outboard)":bump_location_inch
    }
])
car_property_dp = pd.DataFrame([
    {
        "wheelbase (m)":props.wheelbase,
        "trackwidth (m)":props.trackwidth,
        "cg_height (m)":props.cg_height,
        "cp_height (m)":props.cp_height,
        "mass (kg)":props.mass,
        "fwb":props.fwb,
        "C_lA (m^2)":props.cl_a,
        "aero_bal":props.aero_fwb,
        "C_dA (m^2)":props.cd_a,
        "gear_ratio":props.gear_ratio,
        "max_amk_torque (Nm)":props.max_amk_torque,
        "tire_radius (m)":props.tire_radius
    }])

my_ellipse = tire_ellipse(max_accel_g,max_braking_g,max_lateral_g,num_ellipse_points)
pushpoint_ellipse = tire_ellipse(max_accel_g,max_braking_g,max_lateral_g,pushrod_points)
LLTD_range = np.linspace(min_LLTD,max_LLTD,LLTD_points)
bump_range = np.linspace(0,max_bump,bump_amt_points)
bump_location_range = np.linspace(0,bump_location_inch*0.0254,bump_location_points) # assumption! we are not hitting cone left of centerline of CP (right side)
velocity_range = np.linspace(min_v,max_v,v_points)
### call me BROOM the way i sweep ###
big_sweep = pd.DataFrame(
    columns=[
        "a_x","a_y","v","LLTD","bump",
        "FL_Low_Fore","FL_Low_Aft","FL_Up_Fore","FL_Up_Aft","FL_Tie","FL_Push",
        "FR_Low_Fore","FR_Low_Aft","FR_Up_Fore","FR_Up_Aft","FR_Tie","FR_Push",
        "RL_Low_Fore","RL_Low_Aft","RL_Up_Fore","RL_Up_Aft","RL_Toe","RL_Push",
        "RR_Low_Fore","RR_Low_Aft","RR_Up_Fore","RR_Up_Aft","RR_Toe","RR_Push",
        "FL_Fx","FL_Fy","FL_Fz","FL_Mx","FL_My","FL_Mz",
        "FR_Fx","FR_Fy","FR_Fz","FR_Mx","FR_My","FR_Mz",
        "RL_Fx","RL_Fy","RL_Fz","RL_Mx","RL_My","RL_Mz",
        "RR_Fx","RR_Fy","RR_Fz","RR_Mx","RR_My","RR_Mz"
    ]
)
push_sweep = pd.DataFrame(
    columns = [
        "a_x","a_y","v","FR_Push","RR_Push"
    ]
)
for bump in bump_range:
    for lltd in LLTD_range:
        for velocity in velocity_range:
            for bump_location in bump_location_range:
                # update props based on loop variables
                props.v = velocity
                props.drag_force = 0.5 * 1.18 * (velocity**2) * props.cd_a
                props.aero_load = 0.5 * 1.18 * (velocity**2) * props.cl_a
                props.bump_moment_arm_outward = bump_location
                props.LLTD = lltd
                props.bump_amount = max_bump
                sweep_point = real_sweep_tire_ellipse(my_ellipse,masterSus,props)
                push_sweep_point = real_sweep_tire_ellipse(pushpoint_ellipse,masterSus,props)
                for car in sweep_point:
                    # if car.masterCP.a_x<-12.47 and car.masterCP.a_y<-13.65 and car.masterCP.a_x>-12.49 and car.masterCP.a_y>-13.67 and car.props.v>60 and car.props.LLTD>0.54:
                    #     print(car.masterCP.a_x,car.masterCP.a_y,car.props.v,lltd,bump)
                    #     print(car.frCP.contact_patch_forces)                        
                    #     print(str(car.props.drag_force)+" is drag in the car in loop")
                    fl_link_load,fr_link_load,rl_link_load,rr_link_load = get_link_loads_from_car(car)
                    row = {
                        "a_x":car.masterCP.a_x,
                        "a_y":car.masterCP.a_y,
                        "v":velocity,
                        "LLTD":lltd,
                        "bump":bump,
                        "FL_Low_Fore":fl_link_load[0],
                        "FL_Low_Aft":fl_link_load[1],
                        "FL_Up_Fore":fl_link_load[2],
                        "FL_Up_Aft":fl_link_load[3],
                        "FL_Tie":fl_link_load[4],
                        "FL_Push":fl_link_load[5],
                        "FR_Low_Fore":fr_link_load[0],
                        "FR_Low_Aft":fr_link_load[1],
                        "FR_Up_Fore":fr_link_load[2],
                        "FR_Up_Aft":fr_link_load[3],
                        "FR_Tie":fr_link_load[4],
                        "FR_Push":fr_link_load[5],
                        "RL_Low_Fore":rl_link_load[0],
                        "RL_Low_Aft":rl_link_load[1],
                        "RL_Up_Fore":rl_link_load[2],
                        "RL_Up_Aft":rl_link_load[3],
                        "RL_Toe":rl_link_load[4],
                        "RL_Push":rl_link_load[5],
                        "RR_Low_Fore":rr_link_load[0],
                        "RR_Low_Aft":rr_link_load[1],
                        "RR_Up_Fore":rr_link_load[2],
                        "RR_Up_Aft":rr_link_load[3],
                        "RR_Toe":rr_link_load[4],
                        "RR_Push":rr_link_load[5],
                        "FL_Fx":car.flCP.contact_patch_forces[0],
                        "FL_Fy":car.flCP.contact_patch_forces[1],
                        "FL_Fz":car.flCP.contact_patch_forces[2],
                        "FL_Mx":car.flCP.contact_patch_moments[0],
                        "FL_My":car.flCP.contact_patch_moments[1],
                        "FL_Mz":car.flCP.contact_patch_moments[2],
                        "FR_Fx":car.frCP.contact_patch_forces[0],
                        "FR_Fy":car.frCP.contact_patch_forces[1],
                        "FR_Fz":car.frCP.contact_patch_forces[2],
                        "FR_Mx":car.frCP.contact_patch_moments[0],
                        "FR_My":car.frCP.contact_patch_moments[1],
                        "FR_Mz":car.frCP.contact_patch_moments[2],
                        "RL_Fx":car.rlCP.contact_patch_forces[0],
                        "RL_Fy":car.rlCP.contact_patch_forces[1],
                        "RL_Fz":car.rlCP.contact_patch_forces[2],
                        "RL_Mx":car.rlCP.contact_patch_moments[0],
                        "RL_My":car.rlCP.contact_patch_moments[1],
                        "RL_Mz":car.rlCP.contact_patch_moments[2],
                        "RR_Fx":car.rrCP.contact_patch_forces[0],
                        "RR_Fy":car.rrCP.contact_patch_forces[1],
                        "RR_Fz":car.rrCP.contact_patch_forces[2],
                        "RR_Mx":car.rrCP.contact_patch_moments[0],
                        "RR_My":car.rrCP.contact_patch_moments[1],
                        "RR_Mz":car.rrCP.contact_patch_moments[2],
                    }
                    # a_x=-12.48,a_y=-13.66
                    
                    big_sweep = pd.concat([big_sweep,pd.DataFrame([row])],ignore_index=True)
                for car in push_sweep_point:
                    fl_link_load,fr_link_load,rl_link_load,rr_link_load = get_link_loads_from_car(car)
                    push_row = {
                        "a_x":car.masterCP.a_x,
                        "a_y":car.masterCP.a_y,
                        "v":car.props.v,
                        "FL_Push":fl_link_load[5],
                        "FR_Push":fr_link_load[5],
                        "RL_Push":rl_link_load[5],
                        "RR_Push":rr_link_load[5]
                    }
                    push_sweep = pd.concat([push_sweep,pd.DataFrame([push_row])],ignore_index=True)
                if bump==0:
                    break
# Shinyoung Kang
shinyoung = pd.DataFrame()
for ax,ay in pushpoint_ellipse:
    ellipse_point_df = push_sweep[(push_sweep['a_x']==ax) & (push_sweep['a_y']==ay)]
    max_FR_Push = get_max_thing("FR_Push",ellipse_point_df)
    max_RR_Push = get_max_thing("RR_Push",ellipse_point_df)
    max_FL_Push = get_max_thing("FL_Push",ellipse_point_df)
    max_RL_Push = get_max_thing("RL_Push",ellipse_point_df)
    row = {
        "a_x":ax,
        "a_y":ay,
        "FL Push":max_FL_Push,
        "FR Push":max_FR_Push,
        "RL Push":max_RL_Push,
        "RR Push":max_RR_Push,
    }
    shinyoung = pd.concat([shinyoung,pd.DataFrame([row])],ignore_index=True)
signed_max = shinyoung.apply(
    lambda col: col.loc[col.abs().idxmax()] if col.abs().notna().any() else pd.NA
)
shinyoung = pd.concat(
    [signed_max.to_frame().T, shinyoung],
    ignore_index=True
)
# for each ellipse point, get highest magnitude Fx,Fy,Fz Mx,My,Mz on right side
distilled_sweep = pd.DataFrame()
for col in ["a_x","a_y","FR_Fx","FR_Fy","FR_Fz","FR_Mx","FR_My","FR_Mz",
            "RR_Fx","RR_Fy","RR_Fz","RR_Mx","RR_My","RR_Mz","v"]:
    distilled_sweep[col] = big_sweep[col]
max_patch = pd.DataFrame()
for ax,ay in my_ellipse:
    ellipse_point_df = distilled_sweep[(distilled_sweep['a_x']==ax) & (distilled_sweep['a_y']==ay)]
    max_FR_Fx = get_max_thing("FR_Fx",ellipse_point_df)
    max_FR_Fy = get_max_thing("FR_Fy",ellipse_point_df)
    max_FR_Fz = get_max_thing("FR_Fz",ellipse_point_df)
    max_FR_Mx = get_max_thing("FR_Mx",ellipse_point_df)
    max_FR_My = get_max_thing("FR_My",ellipse_point_df)
    max_FR_Mz = get_max_thing("FR_Mz",ellipse_point_df)
    max_RR_Fx = get_max_thing("RR_Fx",ellipse_point_df)
    max_RR_Fy = get_max_thing("RR_Fy",ellipse_point_df)
    max_RR_Fz = get_max_thing("RR_Fz",ellipse_point_df)
    max_RR_Mx = get_max_thing("RR_Mx",ellipse_point_df)
    max_RR_My = get_max_thing("RR_My",ellipse_point_df)
    max_RR_Mz = get_max_thing("RR_Mz",ellipse_point_df)
    row = {
        "a_x":ax,
        "a_y":ay,
        "FR Fx":max_FR_Fx,
        "FR Fy":max_FR_Fy,
        "FR Fz":max_FR_Fz,
        "FR Mx":max_FR_Mx,
        "FR My":max_FR_My,
        "FR Mz":max_FR_Mz,
        "RR Fx":max_RR_Fx,
        "RR Fy":max_RR_Fy,
        "RR Fz":max_RR_Fz,
        "RR Mx":max_RR_Mx,
        "RR My":max_RR_My,
        "RR Mz":max_RR_Mz
    }
    max_patch = pd.concat([max_patch,pd.DataFrame([row])],ignore_index=True)
signed_max = max_patch.apply(
    lambda col: col.loc[col.abs().idxmax()] if col.abs().notna().any() else pd.NA
)
max_patch = pd.concat(
    [signed_max.to_frame().T, max_patch],
    ignore_index=True
)
# Link Extremes
front_tension = [
    max(big_sweep['FL_Low_Fore'].max(),big_sweep['FR_Low_Fore'].max()),
    max(big_sweep['FL_Low_Aft'].max(),big_sweep['FR_Low_Aft'].max()),
    max(big_sweep['FL_Up_Fore'].max(),big_sweep['FR_Up_Fore'].max()),
    max(big_sweep['FL_Up_Aft'].max(),big_sweep['FR_Up_Aft'].max()),
    max(big_sweep['FL_Tie'].max(),big_sweep['FR_Tie'].max()),
    max(big_sweep['FL_Push'].max(),big_sweep['FR_Push'].max())
]
rear_tension = [
    max(big_sweep['RL_Low_Fore'].max(),big_sweep['RR_Low_Fore'].max()),
    max(big_sweep['RL_Low_Aft'].max(),big_sweep['RR_Low_Aft'].max()),
    max(big_sweep['RL_Up_Fore'].max(),big_sweep['RR_Up_Fore'].max()),
    max(big_sweep['RL_Up_Aft'].max(),big_sweep['RR_Up_Aft'].max()),
    max(big_sweep['RL_Toe'].max(),big_sweep['RR_Toe'].max()),
    max(big_sweep['RL_Push'].max(),big_sweep['RR_Push'].max())
]
front_compression = [
    min(big_sweep['FL_Low_Fore'].min(),big_sweep['FR_Low_Fore'].min()),
    min(big_sweep['FL_Low_Aft'].min(),big_sweep['FR_Low_Aft'].min()),
    min(big_sweep['FL_Up_Fore'].min(),big_sweep['FR_Up_Fore'].min()),
    min(big_sweep['FL_Up_Aft'].min(),big_sweep['FR_Up_Aft'].min()),
    min(big_sweep['FL_Tie'].min(),big_sweep['FR_Tie'].min()),
    min(big_sweep['FL_Push'].min(),big_sweep['FR_Push'].min())
]
rear_compression = [
    min(big_sweep['RL_Low_Fore'].min(),big_sweep['RR_Low_Fore'].min()),
    min(big_sweep['RL_Low_Aft'].min(),big_sweep['RR_Low_Aft'].min()),
    min(big_sweep['RL_Up_Fore'].min(),big_sweep['RR_Up_Fore'].min()),
    min(big_sweep['RL_Up_Aft'].min(),big_sweep['RR_Up_Aft'].min()),
    min(big_sweep['RL_Toe'].min(),big_sweep['RR_Toe'].min()),
    min(big_sweep['RL_Push'].min(),big_sweep['RR_Push'].min())
]
front_tension = [round(val,1) for val in front_tension]
rear_tension = [round(val,1) for val in rear_tension]
front_compression = [round(val,1) for val in front_compression]
rear_compression = [round(val,1) for val in rear_compression]
link_extremes = pd.DataFrame({
    "Link":["Low Fore","Low Aft","Up Fore","Up Aft","Tie","Push"],
    "Front Max Tension (N)":front_tension,
    "Front Max Compression (N)":front_compression,
    "Rear Max Tension (N)":rear_tension,
    "Rear Max Compression (N)":rear_compression
})
time = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
sheet_name = f"pebble_master_{time}.xlsx"
with pd.ExcelWriter(sheet_name, engine="openpyxl") as writer:
    big_sweep.to_excel(writer, sheet_name="Master", index=False)
    sim_property_dp.to_excel(writer, sheet_name="Sim_Inputs", index=False)
    car_property_dp.to_excel(writer, sheet_name="Car_Properties", index=False)
    link_extremes.to_excel(writer, sheet_name="Link_Extremes", index=False)
    max_patch.to_excel(writer, sheet_name="Max_Contact_Patch", index=False)
    shinyoung.to_excel(writer, sheet_name="Pushrod_Sweep", index=False)
wb = load_workbook(sheet_name)
#process Master
ws_master = wb["Master"]
num_rows = ws_master.max_row
for row in ws_master["A2":f"B{num_rows}"]:
    for cell in row:
        cell.number_format = '0.00'
for row in ws_master["C2":f"C{num_rows}"]:
    for cell in row:
        cell.number_format = '0.0'
for row in ws_master["D2":f"D{num_rows}"]:
    for cell in row:
        cell.number_format = '0.00'
for row in ws_master["E2":f"E{num_rows}"]:
    for cell in row:
        cell.number_format = '0'
for row in ws_master["F2":f"AZ{num_rows}"]:
    for cell in row:
        cell.number_format = '0'
ws_master.move_range(
    f"A1:{get_column_letter(ws_master.max_column)}{ws_master.max_row}",
    rows=1,
    cols=0
)
ws_master.merge_cells("F1:K1")
ws_master["F1"] = "Front Left"
ws_master.merge_cells("L1:Q1")
ws_master["L1"] = "Front Right"
ws_master.merge_cells("R1:W1")
ws_master["R1"] = "Rear Left"
ws_master.merge_cells("X1:AC1")
ws_master["X1"] = "Rear Right"
links = ["Low Fore","Low Aft","Up Fore","Up Aft","Tie","Push"]
for i,link in enumerate(links):
    fl_col = get_column_letter(6 + i)
    fr_col = get_column_letter(12 + i)
    rl_col = get_column_letter(18 + i)
    rr_col = get_column_letter(24 + i)
    ws_master[f"{fl_col}2"] = f"{link}"
    ws_master[f"{fr_col}2"] = f"{link}"
    ws_master[f"{rl_col}2"] = f"{link}"
    ws_master[f"{rr_col}2"] = f"{link}"
ws_master.merge_cells("A1:A2")
ws_master.merge_cells("B1:B2")
ws_master.merge_cells("C1:C2")
ws_master.merge_cells("D1:D2")
ws_master.merge_cells("E1:E2")
ws_master["A1"] = "a_x (m/s²)"
ws_master["B1"] = "a_y (m/s²)"
ws_master["C1"] = "v (m/s)"
ws_master["D1"] = "LLTD"
ws_master["E1"] = "Bump (N)"
for row in ws_master["A1":f"{get_column_letter(ws_master.max_column)}2"]:
    for cell in row:
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(bold=True)
        cell.border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )

#process Inputs

ws_sim = wb["Sim_Inputs"]
ws_carprops = wb["Car_Properties"]
ws_input = wb.create_sheet("Inputs")
for r_idx, row in enumerate(dataframe_to_rows(sim_property_dp, index=False, header=True), start=1):
    for c_idx,value in enumerate(row, start=1):
        ws_input.cell(row=c_idx, column=r_idx, value=value)
car_prop_start_col = ws_input.max_column + 2
for r_idx, row in enumerate(dataframe_to_rows(car_property_dp, index=False, header=True), start=car_prop_start_col):
    for c_idx, value in enumerate(row, start=1):
        ws_input.cell(row=c_idx, column=r_idx, value=value)
# set A to 22 wide, D to 18
ws_input.column_dimensions["A"].width = 22
ws_input.column_dimensions["D"].width = 18
for row in ws_input["A1":"A9"]:
    for cell in row:
        cell.alignment = Alignment(horizontal="right")
        cell.font = Font(bold=True)
for row in ws_input["D1":"D12"]:
    for cell in row:
        cell.alignment = Alignment(horizontal="right")
        cell.font = Font(bold=True)
#Process Link Extremes
ws_link_extremes = wb["Link_Extremes"]
ws_link_extremes.move_range(
    f"A1:E7",
    rows=1,
    cols=0
)
ws_link_extremes.merge_cells("A1:A2")
ws_link_extremes["A1"] = "Link"
ws_link_extremes.merge_cells("B1:C1")
ws_link_extremes["B1"] = "Front Max"
ws_link_extremes.merge_cells("D1:E1")
ws_link_extremes["D1"] = "Rear Max"
ws_link_extremes["B2"] = "Tension (N)"
ws_link_extremes["C2"] = "Compression (N)"
ws_link_extremes["D2"] = "Tension (N)"
ws_link_extremes["E2"] = "Compression (N)"
ws_link_extremes.column_dimensions["B"].width = 11
ws_link_extremes.column_dimensions["C"].width = 14
ws_link_extremes.column_dimensions["D"].width = 11
ws_link_extremes.column_dimensions["E"].width = 14
for row in ws_link_extremes["A1":"A8"]:
    for cell in row:
        cell.alignment = Alignment(horizontal="right")
for row in ws_link_extremes["A1":"E2"]:
    for cell in row:
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(bold=True)
        cell.border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )
for row in ws_link_extremes["B3":f"E8"]:
    for cell in row:
        cell.number_format = '0'
#process pushrod things
ws_pushrod = wb["Pushrod_Sweep"]
for cell in ws_pushrod[1]:
    cell.font = Font(bold=True)
for row in ws_pushrod["A2":f"B{ws_pushrod.max_row}"]:
    for cell in row:
        cell.number_format = '0.00'
for row in ws_pushrod["C2":f"{get_column_letter(ws_pushrod.max_column)}{ws_pushrod.max_row}"]:
    for cell in row:
        cell.number_format = '0'
ws_pushrod.merge_cells(f"A2:B2")
ws_pushrod["A2"] = "Extremes"
ws_pushrod["A2"].alignment = Alignment(horizontal="right")
for row in ws_pushrod["A1":f"{get_column_letter(ws_pushrod.max_column)}2"]:
    for cell in row:
        cell.font = Font(bold=True)
        cell.border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )
#Max Patch
ws_max_patch = wb["Max_Contact_Patch"]
for cell in ws_max_patch[1]:
    cell.font = Font(bold=True)
for row in ws_max_patch["A2":f"B{ws_max_patch.max_row}"]:
    for cell in row:
        cell.number_format = '0.00'
for row in ws_max_patch["C2":f"{get_column_letter(ws_max_patch.max_column)}{ws_max_patch.max_row}"]:
    for cell in row:
        cell.number_format = '0'
ws_max_patch.merge_cells(f"A2:B2")
ws_max_patch["A2"] = "Extremes"
ws_max_patch["A2"].alignment = Alignment(horizontal="right")
for row in ws_max_patch["A1":f"{get_column_letter(ws_max_patch.max_column)}2"]:
    for cell in row:
        cell.font = Font(bold=True)
        cell.border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )

del wb["Sim_Inputs"]
del wb["Car_Properties"]

wb.save(sheet_name)